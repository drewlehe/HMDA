import gc
import re
import os
import glob
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG
# ============================================================

HMDA_DIR = "/Users/andrew.lehe/Documents/HMDA Data/new_hmdas"
CROSSWALK_PATH = "TRACT_ZIP_122022.xlsx"
OUTPUT_DIR = "/Users/andrew.lehe/Documents/HMDA Data/new_hmda_outputs"

CHUNK_SIZE = 750000

# ============================================================
# HELPERS — post-2017 HMDA LAR format (pipe-delimited, numeric codes)
# ============================================================

def map_occupancy(x) -> str:
    """occupancy_type: 1=principal residence, 2=second residence, 3=investment"""
    try:
        x = int(x)
    except (ValueError, TypeError):
        return "Non-Owner Occupied"
    return "Owner Occupied" if x == 1 else "Non-Owner Occupied"


def map_action(x):
    """action_taken: 1=originated, 2=approved not accepted, 3=denied"""
    try:
        x = int(x)
    except (ValueError, TypeError):
        return None
    if x in {1, 2}:
        return "Approved"
    if x == 3:
        return "Denied"
    return None


def map_purchaser(x) -> str:
    """
    1=Fannie Mae, 2=Ginnie Mae, 3=Freddie Mac,
    6=Commercial bank/savings bank/savings association, everything else=Other
    """
    try:
        x = int(x)
    except (ValueError, TypeError):
        return "Other"
    if x == 6:
        return "Commercial Bank"
    if x == 2:
        return "FHA (Ginnie Mae)"
    if x == 1:
        return "Fannie Mae"
    if x == 3:
        return "Freddie Mac"
    return "Other"


def clean_hmda_chunk(chunk: pd.DataFrame) -> pd.DataFrame:
    chunk["tract_geoid"] = (
        pd.to_numeric(chunk["census_tract"], errors="coerce")
        .astype("Int64")
        .astype(str)
        .str.zfill(11)
    )

    chunk["year"] = pd.to_numeric(chunk["activity_year"], errors="coerce").astype("Int64")

    # loan_amount is in dollars in new format; convert to thousands
    chunk["loan_amount_000s_num"] = pd.to_numeric(chunk["loan_amount"], errors="coerce") / 1000

    # income is already in thousands
    chunk["applicant_income_000s_num"] = pd.to_numeric(chunk["income"], errors="coerce")

    chunk["population_num"] = pd.to_numeric(chunk["tract_population"], errors="coerce")

    # minority_population is a percent; derive count
    chunk["minority_population_num"] = (
        chunk["population_num"]
        * pd.to_numeric(chunk["tract_minority_population_percent"], errors="coerce")
        / 100
    )

    chunk["hud_median_family_income_num"] = pd.to_numeric(
        chunk["ffiec_msa_md_median_family_income"], errors="coerce"
    )
    chunk["tract_to_msamd_income_num"] = pd.to_numeric(
        chunk["tract_to_msa_income_percentage"], errors="coerce"
    )
    chunk["owner_units_num"] = pd.to_numeric(chunk["tract_owner_occupied_units"], errors="coerce")
    chunk["total_units_num"] = pd.to_numeric(chunk["tract_one_to_four_family_homes"], errors="coerce")

    return chunk


def _decode_excel_escapes(s: str) -> str:
    if not isinstance(s, str):
        return s
    return re.sub(r'_x([0-9A-Fa-f]{4})_', lambda m: chr(int(m.group(1), 16)), s)


def load_crosswalk(path: str) -> pd.DataFrame:
    print(f"Loading crosswalk: {path}")
    crosswalk = pd.read_excel(path, engine="openpyxl")
    crosswalk.columns = [str(c).strip().lower() for c in crosswalk.columns]

    required_cols = {"tract", "zip", "tot_ratio"}
    missing = required_cols - set(crosswalk.columns)
    if missing:
        raise ValueError(f"Crosswalk is missing required columns: {sorted(missing)}")

    crosswalk["tract"] = crosswalk["tract"].apply(_decode_excel_escapes)
    crosswalk["zip"] = crosswalk["zip"].apply(_decode_excel_escapes)

    crosswalk["tract_geoid"] = (
        pd.to_numeric(crosswalk["tract"], errors="coerce")
        .astype("Int64")
        .astype(str)
        .str.zfill(11)
    )

    crosswalk["zip_clean"] = (
        pd.to_numeric(crosswalk["zip"], errors="coerce")
        .astype("Int64")
        .astype(str)
        .str.zfill(5)
    )

    for col in ["res_ratio", "bus_ratio", "oth_ratio", "tot_ratio"]:
        if col in crosswalk.columns:
            crosswalk[col] = pd.to_numeric(crosswalk[col], errors="coerce")

    keep_cols = ["tract_geoid", "zip_clean", "tot_ratio"]
    for optional_col in ["res_ratio", "bus_ratio", "oth_ratio"]:
        if optional_col in crosswalk.columns:
            keep_cols.append(optional_col)

    crosswalk = crosswalk.loc[
        crosswalk["tract_geoid"].notna()
        & crosswalk["zip_clean"].notna()
        & crosswalk["tot_ratio"].notna()
    , keep_cols].copy()

    print(f"Crosswalk rows retained: {len(crosswalk):,}")
    return crosswalk


# ============================================================
# EXCEL WRITER HELPERS
# ============================================================

occupancy_order = ["Non-Owner Occupied", "Owner Occupied"]
purchaser_order = ["Commercial Bank", "FHA (Ginnie Mae)", "Fannie Mae", "Freddie Mac", "Other"]

base_cols = [
    "ZIP Code",
    "Population",
    "Minority Population",
    "Median Income",
    "ZIP Code Income/MSA Income",
    "Number of Owner-Occupied Homes",
    "Number of 1-to-4 Family Homes",
]


def build_year_frame(year: int, zip_demo: pd.DataFrame, summary: pd.DataFrame) -> pd.DataFrame:
    demo_y = zip_demo.loc[zip_demo["year"] == year].copy()
    summ_y = summary.loc[summary["year"] == year].copy()

    approved_count = (
        summ_y[["zip_clean", "occupancy_group", "purchaser_group", "decision_group", "mortgage_count"]]
        .rename(columns={"zip_clean": "ZIP Code", "mortgage_count": "value"})
        .assign(metric_name=lambda d: d["decision_group"].map({
            "Approved": "Number of Approved Mortgages",
            "Denied": "Number of Denied Mortgages",
        }))
    )

    approved_amt = (
        summ_y[["zip_clean", "occupancy_group", "purchaser_group", "decision_group", "avg_mortgage_size"]]
        .rename(columns={"zip_clean": "ZIP Code", "avg_mortgage_size": "value"})
        .assign(metric_name=lambda d: d["decision_group"].map({
            "Approved": "Average Approved Mortgage Size",
            "Denied": "Average Denied Mortgage Size",
        }))
    )

    approved_inc = (
        summ_y[["zip_clean", "occupancy_group", "purchaser_group", "decision_group", "avg_applicant_income"]]
        .rename(columns={"zip_clean": "ZIP Code", "avg_applicant_income": "value"})
        .assign(metric_name=lambda d: d["decision_group"].map({
            "Approved": "Average Approved Applicant Income",
            "Denied": "Average Denied Applicant Income",
        }))
    )

    melted = pd.concat([approved_count, approved_amt, approved_inc], ignore_index=True)

    melted["col_key"] = (
        melted["occupancy_group"].astype(str)
        + "|"
        + melted["purchaser_group"].astype(str)
        + "|"
        + melted["metric_name"].astype(str)
    )

    wide = (
        melted.pivot_table(
            index="ZIP Code",
            columns="col_key",
            values="value",
            aggfunc="first",
        )
        .reset_index()
    )

    out = demo_y.merge(wide, on="ZIP Code", how="left")

    ordered_cols = base_cols.copy()
    for occ in occupancy_order:
        for purchaser in purchaser_order:
            ordered_cols.extend([
                f"{occ}|{purchaser}|Number of Approved Mortgages",
                f"{occ}|{purchaser}|Average Approved Mortgage Size",
                f"{occ}|{purchaser}|Average Approved Applicant Income",
                f"{occ}|{purchaser}|Number of Denied Mortgages",
                f"{occ}|{purchaser}|Average Denied Mortgage Size",
                f"{occ}|{purchaser}|Average Denied Applicant Income",
            ])

    for col in ordered_cols:
        if col not in out.columns:
            out[col] = np.nan

    out = out[ordered_cols].sort_values("ZIP Code").reset_index(drop=True)
    return out


def write_sheet(ws, df_year: pd.DataFrame):
    row1, row2, row3 = 1, 2, 3

    thin_gray = Side(style="thin", color="BFBFBF")
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    subheader_fill = PatternFill("solid", fgColor="E2F0D9")
    top_fill = PatternFill("solid", fgColor="BDD7EE")
    bold = Font(bold=True)

    for i, col in enumerate(base_cols, start=1):
        cell = ws.cell(row=row3, column=i, value=col)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin_gray)

    start_col = len(base_cols) + 1
    current_col = start_col

    metric_names = [
        "Number of Approved Mortgages",
        "Average Approved Mortgage Size",
        "Average Approved Applicant Income",
        "Number of Denied Mortgages",
        "Average Denied Mortgage Size",
        "Average Denied Applicant Income",
    ]

    for occ in occupancy_order:
        occ_start = current_col
        for purchaser in purchaser_order:
            purchaser_start = current_col

            for metric in metric_names:
                cell = ws.cell(row=row3, column=current_col, value=metric)
                cell.font = bold
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(bottom=thin_gray)
                current_col += 1

            ws.merge_cells(
                start_row=row2, start_column=purchaser_start,
                end_row=row2, end_column=current_col - 1
            )
            pc = ws.cell(row=row2, column=purchaser_start, value=purchaser)
            pc.font = bold
            pc.fill = subheader_fill
            pc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.merge_cells(
            start_row=row1, start_column=occ_start,
            end_row=row1, end_column=current_col - 1
        )
        oc = ws.cell(row=row1, column=occ_start, value=occ)
        oc.font = Font(bold=True, size=12)
        oc.fill = top_fill
        oc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in [1, 2]:
        for c in range(1, len(base_cols) + 1):
            ws.cell(row=r, column=c).fill = header_fill

    for r_idx, row in enumerate(df_year.itertuples(index=False), start=4):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=None if pd.isna(val) else val)
            col_name = df_year.columns[c_idx - 1]

            if col_name == "ZIP Code":
                cell.number_format = "@"
            elif "Income/MSA Income" in col_name:
                cell.number_format = "0.00"
            elif "Number of " in col_name:
                cell.number_format = "#,##0.00"
            elif "Average " in col_name or col_name in [
                "Median Income", "Population", "Minority Population",
                "Number of Owner-Occupied Homes", "Number of 1-to-4 Family Homes"
            ]:
                cell.number_format = "#,##0.00"

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(ws.max_column)}{ws.max_row}"

    width_map = {"A": 12, "B": 14, "C": 18, "D": 14, "E": 22, "F": 20, "G": 18}
    for col_letter, width in width_map.items():
        ws.column_dimensions[col_letter].width = width
    for c in range(8, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 42


# ============================================================
# PROCESS ONE FILE
# ============================================================

def process_file(hmda_path: str, crosswalk: pd.DataFrame, output_path: str):
    print(f"\n{'='*60}")
    print(f"Processing: {hmda_path}")
    print(f"Output:     {output_path}")
    print(f"{'='*60}")

    tracts_with_zip = set(crosswalk["tract_geoid"].dropna().unique())

    zip_demo_parts = []
    loan_summary_parts = []

    total_rows = 0
    matched_original_rows = 0
    processed_rows_after_match = 0

    reader = pd.read_csv(
        hmda_path,
        sep="|",
        low_memory=False,
        chunksize=CHUNK_SIZE,
        dtype=str,
    )

    for i, chunk in enumerate(reader, start=1):
        print(f"  Chunk {i:,}: {len(chunk):,} rows...")

        total_rows += len(chunk)
        chunk = clean_hmda_chunk(chunk)
        matched_original_rows += chunk["tract_geoid"].isin(tracts_with_zip).sum()

        chunk = chunk.merge(crosswalk, on="tract_geoid", how="left")
        chunk = chunk.loc[chunk["zip_clean"].notna() & chunk["year"].notna()].copy()
        processed_rows_after_match += len(chunk)

        if chunk.empty:
            print("    No matched rows after join/filter.")
            del chunk
            gc.collect()
            continue

        # Per-record mortgage allocations
        chunk["loan_count_alloc"] = chunk["tot_ratio"]
        chunk["loan_amount_alloc_000s"] = chunk["loan_amount_000s_num"] * chunk["tot_ratio"]
        chunk["applicant_income_alloc_000s"] = chunk["applicant_income_000s_num"] * chunk["tot_ratio"]

        chunk["occupancy_group"] = chunk["occupancy_type"].map(map_occupancy)
        chunk["decision_group"] = chunk["action_taken"].map(map_action)
        chunk["purchaser_group"] = chunk["purchaser_type"].map(map_purchaser)

        # Demographic allocations — deduplicate to one row per (tract_geoid, zip_clean)
        # so tract-level attributes are not summed once per HMDA record
        tract_zip = (
            chunk[["year", "tract_geoid", "zip_clean", "tot_ratio",
                   "population_num", "minority_population_num",
                   "hud_median_family_income_num", "tract_to_msamd_income_num",
                   "owner_units_num", "total_units_num"]]
            .drop_duplicates(subset=["tract_geoid", "zip_clean"])
            .copy()
        )
        tract_zip["population_alloc"] = tract_zip["population_num"] * tract_zip["tot_ratio"]
        tract_zip["minority_population_alloc"] = tract_zip["minority_population_num"] * tract_zip["tot_ratio"]
        tract_zip["owner_units_alloc"] = tract_zip["owner_units_num"] * tract_zip["tot_ratio"]
        tract_zip["total_units_alloc"] = tract_zip["total_units_num"] * tract_zip["tot_ratio"]
        tract_zip["income_weight"] = tract_zip["total_units_alloc"]
        tract_zip["msa_income_ratio_weighted_num"] = tract_zip["tract_to_msamd_income_num"] * tract_zip["income_weight"]
        tract_zip["median_income_weighted_num"] = tract_zip["hud_median_family_income_num"] * tract_zip["income_weight"]

        # Keep tract_geoid so we can deduplicate globally across chunks before ZIP-level sum
        demo_part = tract_zip[[
            "year", "tract_geoid", "zip_clean",
            "population_alloc", "minority_population_alloc",
            "owner_units_alloc", "total_units_alloc",
            "msa_income_ratio_weighted_num", "median_income_weighted_num",
            "income_weight",
        ]].copy()
        zip_demo_parts.append(demo_part)

        mortgage_chunk = chunk.loc[chunk["decision_group"].notna()].copy()

        if not mortgage_chunk.empty:
            loan_part = (
                mortgage_chunk.groupby(
                    ["year", "zip_clean", "occupancy_group", "purchaser_group", "decision_group"],
                    dropna=False
                )
                .agg(
                    mortgage_count=("loan_count_alloc", "sum"),
                    loan_amount_sum=("loan_amount_alloc_000s", "sum"),
                    applicant_income_sum=("applicant_income_alloc_000s", "sum"),
                )
                .reset_index()
            )
            loan_summary_parts.append(loan_part)

        del chunk, tract_zip, demo_part, mortgage_chunk
        try:
            del loan_part
        except UnboundLocalError:
            pass
        gc.collect()

    if total_rows == 0:
        print("  WARNING: No rows read. Skipping.")
        return

    coverage = matched_original_rows / total_rows
    print(f"  Total rows: {total_rows:,} | Matched: {matched_original_rows:,} ({coverage:.2%}) | Expanded: {processed_rows_after_match:,}")

    if not zip_demo_parts:
        print("  WARNING: No ZIP-level summaries generated. Skipping.")
        return

    # Global dedup: each (tract_geoid, zip_clean) pair should be counted once,
    # regardless of how many chunks it appeared in
    zip_demo = (
        pd.concat(zip_demo_parts, ignore_index=True)
        .drop_duplicates(subset=["year", "tract_geoid", "zip_clean"])
        .groupby(["year", "zip_clean"], dropna=False)
        .agg(
            Population=("population_alloc", "sum"),
            Minority_Population=("minority_population_alloc", "sum"),
            Owner_Occupied_Homes=("owner_units_alloc", "sum"),
            Total_Homes=("total_units_alloc", "sum"),
            msa_income_ratio_num=("msa_income_ratio_weighted_num", "sum"),
            median_income_num=("median_income_weighted_num", "sum"),
            income_weight_sum=("income_weight", "sum"),
        )
        .reset_index()
    )

    zip_demo["ZIP Code Income/MSA Income"] = zip_demo["msa_income_ratio_num"] / zip_demo["income_weight_sum"]
    zip_demo["Median Income"] = zip_demo["median_income_num"] / zip_demo["income_weight_sum"]

    zip_demo = zip_demo.rename(columns={
        "zip_clean": "ZIP Code",
        "Minority_Population": "Minority Population",
        "Owner_Occupied_Homes": "Number of Owner-Occupied Homes",
        "Total_Homes": "Number of 1-to-4 Family Homes",
    })

    zip_demo = zip_demo[[
        "year", "ZIP Code", "Population", "Minority Population",
        "Median Income", "ZIP Code Income/MSA Income",
        "Number of Owner-Occupied Homes", "Number of 1-to-4 Family Homes",
    ]]

    if loan_summary_parts:
        summary = (
            pd.concat(loan_summary_parts, ignore_index=True)
            .groupby(
                ["year", "zip_clean", "occupancy_group", "purchaser_group", "decision_group"],
                dropna=False
            )
            .agg(
                mortgage_count=("mortgage_count", "sum"),
                loan_amount_sum=("loan_amount_sum", "sum"),
                applicant_income_sum=("applicant_income_sum", "sum"),
            )
            .reset_index()
        )
        summary["avg_mortgage_size"] = summary["loan_amount_sum"] / summary["mortgage_count"]
        summary["avg_applicant_income"] = summary["applicant_income_sum"] / summary["mortgage_count"]
    else:
        summary = pd.DataFrame(columns=[
            "year", "zip_clean", "occupancy_group", "purchaser_group", "decision_group",
            "mortgage_count", "loan_amount_sum", "applicant_income_sum",
            "avg_mortgage_size", "avg_applicant_income"
        ])

    summary["zip_clean"] = summary["zip_clean"].astype(str)

    years = sorted(pd.to_numeric(zip_demo["year"], errors="coerce").dropna().astype(int).unique().tolist())
    print(f"  Years found: {years}")

    # ---- Sanity checks ----
    print("  --- Demographic sanity check ---")
    print(f"  ZIP count: {len(zip_demo):,}")
    for col in ["Population", "Minority Population", "Median Income",
                "ZIP Code Income/MSA Income", "Number of Owner-Occupied Homes", "Number of 1-to-4 Family Homes"]:
        s = zip_demo[col]
        print(f"  {col}: min={s.min():,.1f}  median={s.median():,.1f}  max={s.max():,.1f}  nulls={s.isna().sum()}")
    pop_violations = (zip_demo["Minority Population"] > zip_demo["Population"]).sum()
    home_violations = (zip_demo["Number of Owner-Occupied Homes"] > zip_demo["Number of 1-to-4 Family Homes"]).sum()
    print(f"  Minority > Population violations: {pop_violations}")
    print(f"  Owner Homes > Total Homes violations: {home_violations}")
    if not summary.empty:
        approved = summary.loc[summary["decision_group"] == "Approved", "mortgage_count"].sum()
        denied   = summary.loc[summary["decision_group"] == "Denied",   "mortgage_count"].sum()
        print(f"  Total approved mortgages (weighted): {approved:,.0f}")
        print(f"  Total denied mortgages  (weighted): {denied:,.0f}")
        print(f"  Avg mortgage size (approved): {(summary.loc[summary['decision_group']=='Approved','avg_mortgage_size'].mean()):,.0f}k")
    print("  --- End sanity check ---")

    print("  Writing Excel workbook...")
    wb = Workbook()
    wb.remove(wb.active)

    for year in years:
        df_year = build_year_frame(year, zip_demo, summary)
        ws = wb.create_sheet(title=str(year))
        write_sheet(ws, df_year)

    wb.save(output_path)
    print(f"  Saved: {output_path}")


# ============================================================
# MAIN
# ============================================================

def main():
    crosswalk = load_crosswalk(CROSSWALK_PATH)

    pattern = os.path.join(HMDA_DIR, "*_lar.txt")
    files = sorted(glob.glob(pattern))

    if not files:
        raise FileNotFoundError(f"No LAR files found matching: {pattern}")

    print(f"\nFound {len(files)} file(s) to process:")
    for f in files:
        print(f"  {os.path.basename(f)}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    for hmda_path in files:
        basename = os.path.basename(hmda_path)
        year_match = re.search(r'(\d{4})_lar', basename)
        year_str = year_match.group(1) if year_match else "unknown"
        output_path = os.path.join(OUTPUT_DIR, f"hmda_{year_str}_zip_summary_weighted_chunked.xlsx")
        process_file(hmda_path, crosswalk, output_path)

    print("\nAll files processed.")


if __name__ == "__main__":
    main()
